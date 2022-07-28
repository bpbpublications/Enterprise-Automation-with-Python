from openpyxl import load_workbook

workbook = load_workbook(filename="test_workbook.xlsx")
sheet = workbook.active

# Using iterator to get rows values
for value in sheet.iter_rows(
    min_row=1, max_row=2, min_col=1, max_col=3
):
    print("Rows Values:" + str(value))

