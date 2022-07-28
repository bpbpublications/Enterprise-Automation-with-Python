from openpyxl import load_workbook

workbook = load_workbook(filename="test_workbook.xlsx")
sheet = workbook.active

# functions to print data from the active sheet
def print_data():
     for row in sheet.iter_rows(values_only=True):
         print(row)

sheet["A2"] = "Adding more data"
sheet["A3"] = "Adding more data"
sheet["B3"] = "Adding more data"

print_data()

# Save the workbook
workbook.save(filename="test_workbook.xlsx")
