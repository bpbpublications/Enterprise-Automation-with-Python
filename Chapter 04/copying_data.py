from openpyxl import load_workbook

# opening the source excel file
source_file ="source_workbook.xlsx"
source_workbook = load_workbook(source_file)
source_worksheet = source_workbook.worksheets[0]
# opening the destination excel file
destination_file ="destination_workbook.xlsx"
destination_workbook = load_workbook(destination_file)
destination_worksheet = destination_workbook.active
# calculate total number of rows in source excel file
max_row = source_worksheet.max_row
# copying the column 1 and 2
for row in range (1, max_row + 1):
    for column in range (1, 3):
        # reading cell value from source excel file
        c = source_worksheet.cell(row = row, column = column)
        # writing the read value to destination excel file
        destination_worksheet.cell(row = row, column = column).value = c.value
# saving the destination excel file
destination_workbook.save(destination_file)
