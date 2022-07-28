from openpyxl import load_workbook

workbook = load_workbook(filename="test_workbook.xlsx")
sheet = workbook.active

# Looping through row and its values
for rows in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=3):
    for row in rows:
        print("Row is: " + str(row) + " and value is: " + str(row.value))
