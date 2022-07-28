from openpyxl import load_workbook

# load the test workbook
workbook = load_workbook(filename="test_workbook.xlsx")

# Note the str() function is used to convert the return type to string
print("Sheetnames are: " + str(workbook.sheetnames))
sheet = workbook.active

# Access sheet element value by id
print("Value of cell A1: " + sheet["A1"].value)
# Access sheet element value by row column id
print("Value of cell with row as 1 and column as 2 is: " + str(sheet.cell(row=1, column=2).value))
